from flask import Blueprint, render_template, request, send_file, redirect, url_for
from flask_login import login_required, current_user
from models import db, Employee, Payroll
from datetime import datetime, date, timedelta
from functools import wraps
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

reports_bp = Blueprint('reports', __name__, url_prefix='/reports')

def admin_or_hr_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if current_user.role not in ['admin', 'hr']:
            return redirect(url_for('dashboard.index'))
        return f(*args, **kwargs)
    return decorated_function

@reports_bp.route('/')
@login_required
@admin_or_hr_required
def list_reports():
    """Página de reportes"""
    return render_template('reports/index.html')

@reports_bp.route('/monthly', methods=['GET', 'POST'])
@login_required
@admin_or_hr_required
def monthly_report():
    """Reporte mensual de nóminas"""
    year = request.args.get('year', datetime.now().year, type=int)
    month = request.args.get('month', datetime.now().month, type=int)
    
    # Generar fecha de inicio y fin
    first_day = date(year, month, 1)
    if month == 12:
        last_day = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = date(year, month + 1, 1) - timedelta(days=1)
    
    # Buscar nóminas del período
    payrolls = Payroll.query.filter(
        Payroll.period_start >= first_day,
        Payroll.period_end <= last_day
    ).all()
    
    # Calcular totales
    total_gross = sum(p.gross_salary for p in payrolls)
    total_discounts = sum(p.total_discounts for p in payrolls)
    total_net = sum(p.net_salary for p in payrolls)
    total_tax = sum(p.income_tax for p in payrolls)
    total_afp = sum(p.afp for p in payrolls)
    total_insurance = sum(p.insurance for p in payrolls)
    
    context = {
        'year': year,
        'month': month,
        'payrolls': payrolls,
        'total_gross': total_gross,
        'total_discounts': total_discounts,
        'total_net': total_net,
        'total_tax': total_tax,
        'total_afp': total_afp,
        'total_insurance': total_insurance,
    }
    
    return render_template('reports/monthly.html', **context)

@reports_bp.route('/employee/<int:employee_id>')
@login_required
def employee_report(employee_id):
    """Reporte por empleado"""
    employee = Employee.query.get_or_404(employee_id)
    payrolls = employee.payrolls
    
    context = {
        'employee': employee,
        'payrolls': payrolls,
    }
    
    return render_template('reports/employee.html', **context)

@reports_bp.route('/export/monthly')
@login_required
@admin_or_hr_required
def export_monthly_excel():
    """Exportar reporte mensual a Excel"""
    year = request.args.get('year', datetime.now().year, type=int)
    month = request.args.get('month', datetime.now().month, type=int)
    
    first_day = date(year, month, 1)
    if month == 12:
        last_day = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = date(year, month + 1, 1) - timedelta(days=1)
    
    payrolls = Payroll.query.filter(
        Payroll.period_start >= first_day,
        Payroll.period_end <= last_day
    ).all()
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nómina"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1f4788", end_color="1f4788", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Encabezados
    headers = ["Empleado", "Cédula", "Salario Base", "Horas Extras", "Bonificaciones", 
               "Bruto", "Impuesto", "AFP", "Seguro", "Otros Desc.", "Total Desc.", "Neto"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # Datos
    for row, payroll in enumerate(payrolls, 2):
        ws.cell(row=row, column=1, value=payroll.employee.full_name)
        ws.cell(row=row, column=2, value=payroll.employee.id_number)
        ws.cell(row=row, column=3, value=payroll.base_salary)
        ws.cell(row=row, column=4, value=payroll.overtime_pay)
        ws.cell(row=row, column=5, value=payroll.bonuses)
        ws.cell(row=row, column=6, value=payroll.gross_salary)
        ws.cell(row=row, column=7, value=payroll.income_tax)
        ws.cell(row=row, column=8, value=payroll.afp)
        ws.cell(row=row, column=9, value=payroll.insurance)
        ws.cell(row=row, column=10, value=payroll.other_discounts)
        ws.cell(row=row, column=11, value=payroll.total_discounts)
        ws.cell(row=row, column=12, value=payroll.net_salary)
    
    # Ancho de columnas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"nomina_{year}_{month:02d}.xlsx"
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
